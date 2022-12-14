import csv
import itertools
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from typing import Dict, Tuple, List, Callable


class Vacancy:
    def __init__(self, dict_vac: Dict[str, str]):
        self.name: str = dict_vac['name']
        self.salary = self.get_medium_salary(dict_vac['salary_from'], dict_vac['salary_to'],
                                             dict_vac['salary_currency'])
        self.area_name = dict_vac['area_name']
        self.year: int = int(dict_vac['published_at'][:4])

    def get_medium_salary(self, salary_from: str, salary_to: str, salary_currency: str):
        salary_from = salary_from.split('.')[0]
        salary_to = salary_to.split('.')[0]
        raw_currency = salary_currency
        medium = (int(salary_from) + int(salary_to)) / 2
        return medium * self.__currency_to_rub[raw_currency]

    __currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }


class DataSet:
    def __init__(self, file_name: str, prof_name: str):
        list_raw_vacancies = self.__csv_filer(*self.__csv_reader(file_name))
        self.__list_vacs = [Vacancy(vacancy) for vacancy in list_raw_vacancies]

        grouped_by_year = self.group_by_year()
        grouped_by_name = self.group_by_year_with_name(prof_name)
        self.prof_name = prof_name
        self.years = list(grouped_by_year.keys())
        self.salary_by_year_dict = self.salary_by_years(grouped_by_year)
        self.count_by_year_dict = self.count_by_years(grouped_by_year)
        self.salary_by_year_name_dict = self.salary_by_years(grouped_by_name)
        self.count_by_year_name_dict = self.count_by_years(grouped_by_name)
        grouped_by_city, dola_group = self.group_by_city()
        count_first_cities = 10
        self.salary_by_city_dict = dict(
            itertools.islice(self.salary_by_city(grouped_by_city).items(), count_first_cities))
        self.percent_by_city_dict = dict(itertools.islice(dola_group.items(), count_first_cities))
        self.cities_by_salary = list(self.salary_by_city_dict.keys())
        self.cities_by_percent = list(self.percent_by_city_dict.keys())

    def __csv_reader(self, file_name: str) -> Tuple[List, List]:
        with open(file_name, newline='') as csvfile:
            reader = csv.reader(csvfile, delimiter=',')
            data = [line for line in reader]
            return data[0], data[1:]

    def __csv_filer(self, headers: List[str], data: List[str]) -> List[Dict[str, str]]:
        array = []
        for vacancy in data:
            categories = [category for category in vacancy if len(category) != 0]
            if len(categories) != len(headers):
                continue
            array.append(dict(zip(headers, categories)))
        return array

    @staticmethod
    def clean_text(text: str) -> str:
        cleaned_text = re.sub(re.compile('<.*?>'), '', text).strip()
        return ';'.join(cleaned_text.split('\n')) if '\n' in cleaned_text else ' '.join(cleaned_text.split())

    def group_by_year(self) -> Dict[int, List[Vacancy]]:
        dict_years = {}
        for vacancy in self.__list_vacs:
            if not dict_years.keys().__contains__(vacancy.year):
                dict_years[vacancy.year] = [vacancy]
            else:
                dict_years[vacancy.year].append(vacancy)
        return dict_years

    def group_by_year_with_name(self, name: str) -> Dict[int, List[Vacancy]]:
        dict_years = {}
        for vacancy in self.__list_vacs:
            if name in vacancy.name:
                if not dict_years.keys().__contains__(vacancy.year):
                    dict_years[vacancy.year] = [vacancy]
                else:
                    dict_years[vacancy.year].append(vacancy)
        return dict_years

    def group_by_city(self) -> Tuple[Dict[str, List[Vacancy]], Dict[str, float]]:
        dict_city = {}
        for vacancy in self.__list_vacs:
            if not dict_city.keys().__contains__(vacancy.area_name):
                dict_city[vacancy.area_name] = [vacancy]
            else:
                dict_city[vacancy.area_name].append(vacancy)
        dict_city_more_1percent = {}
        fraction_by_city = {}
        for city, list_vac in dict_city.items():
            fraction = len(dict_city[city]) / len(self.__list_vacs)
            if fraction * 100 >= 1:
                dict_city_more_1percent[city] = list_vac
                fraction_by_city[city] = round(fraction, 4)
        return dict_city_more_1percent, dict(sorted(fraction_by_city.items(), key=lambda x: x[1], reverse=True))

    def salary_by_years(self, dict_vacs: Dict[int, List[Vacancy]]):
        ready_dict = {}
        for year in self.years:
            if dict_vacs.keys().__contains__(year):
                ready_dict[year] = int(sum([vac.salary for vac in dict_vacs[year]]) / len(dict_vacs[year]))
            else:
                ready_dict[year] = 0
        return ready_dict

    def salary_by_city(self, dict_vacs: Dict[str, List[Vacancy]]):
        salary_by_city_dict = {}
        for city, list_vacancy in dict_vacs.items():
            salary_by_city_dict[city] = int(sum([vac.salary for vac in list_vacancy]) / len(list_vacancy))
        return dict(sorted(salary_by_city_dict.items(), key=lambda x: x[1], reverse=True))

    def count_by_years(self, _dict: Dict[int, List[Vacancy]]):
        count_by_years_dict = {}
        for year in self.years:
            count_by_years_dict[year] = len(_dict[year]) if _dict.keys().__contains__(year) else 0
        return count_by_years_dict

    def print(self):
        print('Динамика уровня зарплат по годам: ' + str(self.salary_by_year_dict))
        print('Динамика количества вакансий по годам: ' + str(self.count_by_year_dict))
        print('Динамика уровня зарплат по годам для выбранной профессии: ' + str(
            self.salary_by_year_name_dict))
        print('Динамика количества вакансий по годам для выбранной профессии: ' + str(
            self.count_by_year_name_dict))
        print('Уровень зарплат по городам (в порядке убывания): ' + str(self.salary_by_city_dict))
        print('Доля вакансий по городам (в порядке убывания): ' + str(self.percent_by_city_dict))


class Report:

    def __init__(self, data_set: DataSet):
        self.__data_set = data_set

    def generate_excel(self):
        wb = Workbook()
        by_year = wb.active
        by_year.title = "Статистика по годам"
        by_city = wb.create_sheet("Статистика по городам")

        data_set = self.__data_set
        self.__fill_sheet(["Год", "Средняя зарплата", f"Средняя зарплата - {self.__data_set.prof_name}",
                           "Количество вакансий", f"Количество вакансий - {self.__data_set.prof_name}"],
                          lambda i: [data_set.years[i], data_set.salary_by_year_dict[data_set.years[i]],
                                     data_set.salary_by_year_name_dict[data_set.years[i]],
                                     data_set.count_by_year_dict[data_set.years[i]],
                                     data_set.count_by_year_name_dict[data_set.years[i]]],
                          by_year, len(data_set.years))

        self.__fill_sheet(["Город", "Уровень зарплат", "",
                           "Город", "Доля вакансий"], lambda i: [data_set.cities_by_salary[i],
                                                                 data_set.salary_by_city_dict[
                                                                     data_set.cities_by_salary[i]], '',
                                                                 data_set.cities_by_percent[i],
                                                                 data_set.percent_by_city_dict[
                                                                     data_set.cities_by_percent[i]]], by_city,
                          len(data_set.cities_by_salary))
        self.__apply_styles(by_city)
        self.__apply_styles(by_year)

        wb.save('report.xlsx')

    def __apply_styles(self, ws):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for cell in next(ws.rows):
            cell.font = Font(bold=True)
        for column_cells in ws.columns:
            padding = 3
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + padding
            for cell in column_cells:
                cell.border = thin_border

    def __fill_sheet(self, header: List[str], x: Callable[[int], List], ws, count_row: int):
        self.fill_row(1, header, ws)
        for index in range(0, count_row):
            self.fill_row(index + 2, x(index), ws)

    def fill_row(self, row_index: int, data: List, ws):
        for column_index in range(0, len(data)):
            ws[get_column_letter(column_index + 1) + str(row_index)] = data[column_index]


file_name = input('Введите название файла: ')
name = input('Введите название профессии: ')
data = DataSet(file_name, name)
Report(data).generate_excel()