import csv
import itertools
import re
import cProfile
import numpy as np
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from typing import Dict, Tuple, List, Callable

pr = cProfile.Profile()
pr.enable()

class Vacancy:
    """Класс вакансии, хранящий информацию о конкретной вакансии.

    """

    def __init__(self, dict_vac: Dict[str, str]):
        """Инициализирует обхект Vacancy.

        :param dict_vac: Массиы сырых данных.
        """
        self.name: str = dict_vac['name']
        self.salary = self.get_medium_salary(dict_vac['salary_from'], dict_vac['salary_to'],
                                             dict_vac['salary_currency'])
        self.area_name = dict_vac['area_name']
        self.year: int = int(dict_vac['published_at'][:4])

    def get_medium_salary(self, salary_from: str, salary_to: str, salary_currency: str):
        """Метод получения средней зарплаты в рублях

        :param salary_from: Нижняя граница оклада.
        :param salary_to: Верхняя граница оклада.
        :param salary_currency: Валюта оклада.
        :return: Среднюю оклада.
        """
        salary_from = salary_from.split('.')[0]
        salary_to = salary_to.split('.')[0]
        raw_currency = salary_currency
        medium = (int(salary_from) + int(salary_to)) / 2
        return medium * self.__currency_to_rub[raw_currency]

    __currency_to_rub: dict[str, float | int] = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }


class DataSet:
    """Класс для обработки csv файлов

    """
    def __init__(self, file_name: str, prof_name: str):
        """Инициализирует объект DataSet

        :param file_name: Имя файла.
        :param prof_name: Название профессии, по которой требуется более подробная информация.
        """
        list_raw_vacancies = self.__csv_filer(*self.__csv_reader(file_name))
        self.__list_vacs = [Vacancy(vacancy) for vacancy in list_raw_vacancies]

        grouped_by_year = self.group_by_year()
        grouped_by_name = self.group_by_year_with_name(prof_name)
        self.prof_name = prof_name
        self.years = list(grouped_by_year.keys())
        self.salary_by_year_dict = self.salary_by_years(grouped_by_year)
        self.count_by_year_dict = self.count_by_years(grouped_by_year)
        self.salary_by_year_name_dict = self.salary_by_years(grouped_by_name)
        self.count_by_year_name_dict = self.count_by_years(grouped_by_name)
        grouped_by_city, dola_group = self.group_by_city()
        count_first_cities = 10
        self.salary_by_city_dict = dict(
            itertools.islice(self.salary_by_city(grouped_by_city).items(), count_first_cities))
        self.percent_by_city_dict = dict(itertools.islice(dola_group.items(), count_first_cities))
        self.cities_by_salary = list(self.salary_by_city_dict.keys())
        self.cities_by_percent = list(self.percent_by_city_dict.keys())

    def __csv_reader(self, file_name: str) -> Tuple[List, List]:
        """Читает csv файл.

        :param file_name: Имя файла.
        :return: Tuple из списка заголовков и данных.
        """
        with open(file_name, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=',')
            data = [line for line in reader]
            return data[0], data[1:]

    def __csv_filer(self, headers: List[str], data: List[str]) -> List[Dict[str, str]]:
        """Преобразует массив сырых данных в массив словарей.

        :param headers: Массив заголовков.
        :param data: Массив сырых данных.
        :return: Массив словарей, где ключ - название столбца.
        """
        array = []
        for vacancy in data:
            categories = [category for category in vacancy if len(category) != 0]
            if len(categories) != len(headers):
                continue
            array.append(dict(zip(headers, categories)))
        return array

    @staticmethod
    def clean_text(text: str) -> str:
        """Очищает текст от лишних символов.

        :param text: Сырой текст.
        :return: Чистый текст.
        """
        cleaned_text = re.sub(re.compile('<.*?>'), '', text).strip()
        return ';'.join(cleaned_text.split('\n')) if '\n' in cleaned_text else ' '.join(cleaned_text.split())

    def group_by_year(self) -> Dict[int, List[Vacancy]]:
        """Группирует массивы вакансий по годам.

        :return: Словарь {год: массив вакнсий}.
        """
        dict_years = {}
        for vacancy in self.__list_vacs:
            if not dict_years.keys().__contains__(vacancy.year):
                dict_years[vacancy.year] = [vacancy]
            else:
                dict_years[vacancy.year].append(vacancy)
        return dict_years

    def group_by_year_with_name(self, name: str) -> Dict[int, List[Vacancy]]:
        """Группирует массивы вакансий по годам только для конкретной профессии.

        :param name: Название профессии.
        :return: Словарь {год: массив вакнсий}.
        """
        dict_years = {}
        for vacancy in self.__list_vacs:
            if name in vacancy.name:
                if not dict_years.keys().__contains__(vacancy.year):
                    dict_years[vacancy.year] = [vacancy]
                else:
                    dict_years[vacancy.year].append(vacancy)
        return dict_years

    def group_by_city(self) -> Tuple[Dict[str, List[Vacancy]], Dict[str, float]]:
        """Группирует массивы вакансий по городам.

        :return: Словарь {город: массив вакнсий}.
        """
        dict_city = {}
        for vacancy in self.__list_vacs:
            if not dict_city.keys().__contains__(vacancy.area_name):
                dict_city[vacancy.area_name] = [vacancy]
            else:
                dict_city[vacancy.area_name].append(vacancy)
        dict_city_more_1percent = {}
        fraction_by_city = {}
        for city, list_vac in dict_city.items():
            fraction = len(dict_city[city]) / len(self.__list_vacs)
            if fraction * 100 >= 1:
                dict_city_more_1percent[city] = list_vac
                fraction_by_city[city] = round(fraction, 4)
        return dict_city_more_1percent, dict(sorted(fraction_by_city.items(), key=lambda x: x[1], reverse=True))

    def salary_by_years(self, dict_vacs: Dict[int, List[Vacancy]]) -> Dict[int, float]:
        """Группирует массив ЗП по годам.

        :param dict_vacs: Словарь вакансий по годам.
        :return: Словарь {год: средняя ЗП}
        """
        ready_dict = {}
        for year in self.years:
            if dict_vacs.keys().__contains__(year):
                ready_dict[year] = int(sum([vac.salary for vac in dict_vacs[year]]) / len(dict_vacs[year]))
            else:
                ready_dict[year] = 0
        return ready_dict

    def salary_by_city(self, dict_vacs: Dict[str, List[Vacancy]]) -> Dict[str, float]:
        """Группирует ЗП по городам.

        :param dict_vacs: Словарь вакансий по городам.
        :return: Словарь {город: средняя ЗП}
        """
        salary_by_city_dict = {}
        for city, list_vacancy in dict_vacs.items():
            salary_by_city_dict[city] = int(sum([vac.salary for vac in list_vacancy]) / len(list_vacancy))
        return dict(sorted(salary_by_city_dict.items(), key=lambda x: x[1], reverse=True))

    def count_by_years(self, _dict: Dict[int, List[Vacancy]]) -> Dict[int, int]:
        """Группирует количество вакансий по годам.

        :param _dict: Словарь {год: массив вакансий}.
        :return: Словарь {год: количество вакансий}.
        """
        count_by_years_dict = {}
        for year in self.years:
            count_by_years_dict[year] = len(_dict[year]) if _dict.keys().__contains__(year) else 0
        return count_by_years_dict

    def print(self):
        """Выводит в консоль всю обработанную информацию.

        """
        print('Динамика уровня зарплат по годам: ' + str(self.salary_by_year_dict))
        print('Динамика количества вакансий по годам: ' + str(self.count_by_year_dict))
        print('Динамика уровня зарплат по годам для выбранной профессии: ' + str(
            self.salary_by_year_name_dict))
        print('Динамика количества вакансий по годам для выбранной профессии: ' + str(
            self.count_by_year_name_dict))
        print('Уровень зарплат по городам (в порядке убывания): ' + str(self.salary_by_city_dict))
        print('Доля вакансий по городам (в порядке убывания): ' + str(self.percent_by_city_dict))


class ReportGraphic:
    """Класс для создания PNG картинки с графиками, исходя из данных ему данных.

    """

    def __init__(self, data_set: DataSet):
        """Инициализирует ReportGraphic.

        :param data_set: Готовые данные для вывода.
        """
        self.__data_set = data_set

    def generate_image(self):
        """Метод генерации итогового изображения.

        """
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)
        self.generate_bar(ax=ax1,
                          axis_x=self.__data_set.years,
                          axes_y=[self.__data_set.salary_by_year_dict.values(),
                                  self.__data_set.salary_by_year_name_dict.values()],
                          title='Уровень зарплат по годам',
                          labels=['Средняя з/п', f'з/п {self.__data_set.prof_name.lower()}'],
                          width=0.4)
        self.generate_bar(ax=ax2,
                          axis_x=self.__data_set.years,
                          axes_y=[self.__data_set.count_by_year_dict.values(),
                                  self.__data_set.count_by_year_name_dict.values()],
                          title='Количество вакансий по годам',
                          labels=['Количество вакансий', f'Количество вакансий \n{self.__data_set.prof_name.lower()}'],
                          width=0.4)
        cities_by_salary = ['\n'.join(re.split(r'-| ', city_name)) for city_name in self.__data_set.cities_by_salary]
        self.generate_barh(ax=ax3,
                           axis_x=cities_by_salary,
                           axes_y=list(self.__data_set.salary_by_city_dict.values()),
                           title='Уровень зарплат по городам')
        self.generate_pie(ax=ax4,
                          date=self.__data_set.percent_by_city_dict.values(),
                          labels=self.__data_set.cities_by_percent,
                          title='Доля вакансий по городам')

        plt.tight_layout()
        plt.savefig('graph.png')

    def generate_bar(self, ax, axis_x: List[int], axes_y: List[List[int]], title: str, labels: List[str],
                     width: float):
        """Генерирует фигуру типа bar.

        :param ax: Часть холста.
        :param axis_x: Данные для оси X на графике.
        :param axes_y: Данные для оси Y на графике.
        :param title: Заголовок графика.
        :param labels: Подписи к оси X.
        :param width: Ширина столбца.
        """
        if len(axes_y) != len(labels):
            raise Exception('Неодинаковая длина labes и axes_y')
        fontsize = 8
        x = np.arange(axis_x[0], axis_x[len(axis_x) - 1] + 1)
        for index in range(len(axes_y)):
            ax.bar(x - width / 2 if index % 2 == 0 else x + width / 2,
                   axes_y[index],
                   width=width,
                   label=labels[index])
        ax.set_title(title)
        ax.set_xticks(x)
        ax.set_xticklabels(x, rotation=90, fontsize=fontsize)
        ax.grid(axis='y')
        ax.legend(fontsize=fontsize)

    def generate_barh(self, ax, axis_x: List[str], axes_y: List[int], title: str):
        """Генерирует горизонтальный Bar.

        :param ax:  Часть холста.
        :param axis_x: Данные для оси X на графике.
        :param axes_y: Данные для оси Y на графике.
        :param title: Заголовок графика.
        """
        fontsize = 6
        ax.barh(axis_x, axes_y)
        ax.set_title(title)
        ax.invert_yaxis()
        ax.set_yticks(axis_x)
        ax.set_yticklabels(axis_x, fontsize=fontsize)
        ax.grid(axis='x')

    def generate_pie(self, ax, date: List[int], title: str, labels=List[str]):
        """Генерация графика типа PIE(типо пирог:))

        :param ax:
        :param date:
        :param title:
        :param labels:
        """
        fontsize = 6
        ax.pie(date, labels=labels, textprops={'fontsize': fontsize})
        ax.set_title(title)
        ax.grid(axis='x')


class ReportTable:
    """Генерирует EXCEL таблицу из данных класса DataSet"""

    def __init__(self, data_set: DataSet):
        """Инициализирует объект ReportTable.

        :param data_set: Готовые данные для вывода.
        """
        self.__data_set = data_set

    def generate_excel(self):
        """Генерирует EXCEL табличку.

        """
        wb = Workbook()
        by_year = wb.active
        by_year.title = "Статистика по годам"
        by_city = wb.create_sheet("Статистика по городам")

        data_set = self.__data_set
        self.__fill_sheet(["Год", "Средняя зарплата", f"Средняя зарплата - {self.__data_set.prof_name}",
                           "Количество вакансий", f"Количество вакансий - {self.__data_set.prof_name}"],
                          lambda i: [data_set.years[i], data_set.salary_by_year_dict[data_set.years[i]],
                                     data_set.salary_by_year_name_dict[data_set.years[i]],
                                     data_set.count_by_year_dict[data_set.years[i]],
                                     data_set.count_by_year_name_dict[data_set.years[i]]],
                          by_year, len(data_set.years))

        self.__fill_sheet(["Город", "Уровень зарплат", "",
                           "Город", "Доля вакансий"], lambda i: [data_set.cities_by_salary[i],
                                                                 data_set.salary_by_city_dict[
                                                                     data_set.cities_by_salary[i]], '',
                                                                 data_set.cities_by_percent[i],
                                                                 data_set.percent_by_city_dict[
                                                                     data_set.cities_by_percent[i]]], by_city,
                          len(data_set.cities_by_salary))
        self.__apply_styles(by_city)
        self.__apply_styles(by_year)

        wb.save('report.xlsx')

    def __apply_styles(self, ws):
        """Применяет стили для листа.

        :param ws: Лист таблицы.
        """
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for cell in next(ws.rows):
            cell.font = Font(bold=True)
        for column_cells in ws.columns:
            padding = 3
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + padding
            for cell in column_cells:
                cell.border = thin_border

    def __fill_sheet(self, header: List[str], x: Callable[[int], List], ws, count_row: int):
        """Заполняет лист данными.

        :param header: Заголовки для листа.
        :param x: Лямбда функция, возвращающая массив для заполнения строки по индексу.
        :param ws: Лист таблицы.
        :param count_row: Количество строк.
        """
        self.fill_row(1, header, ws)
        for index in range(0, count_row):
            self.fill_row(index + 2, x(index), ws)

    def fill_row(self, row_index: int, data: List, ws):
        """Заполняет строку листа.

        :param row_index: Индекс строки для заполнения.
        :param data: Информация для заполения.
        :param ws: Лист таблицы.
        """
        for column_index in range(0, len(data)):
            ws[get_column_letter(column_index + 1) + str(row_index)] = data[column_index]


report_type = input('Вакансии или Статистика: ')
file_name = input('Введите название файла: ')
name = input('Введите название профессии: ')
if report_type == 'Вакансии':
    data = DataSet(file_name, name)
    ReportTable(data).generate_excel()
else:
    data = DataSet(file_name, name)
    ReportGraphic(data).generate_image()
pr.disable()
pr.print_stats()